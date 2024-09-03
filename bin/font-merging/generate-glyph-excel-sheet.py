from argparse import ArgumentParser
from pathlib import Path
from typing import Tuple, List
from hashlib import md5
from fontTools.ttLib.ttFont import TTFont
from fontTools.ttLib.tables._g_l_y_f import Glyph
from fontTools.pens.recordingPen import DecomposingRecordingPen
from fontTools.pens.freetypePen import FreeTypePen
from fontTools.merge.cmap import _glyphsAreSame
from openpyxl import Workbook
import openpyxl.drawing.image
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
from freetype import ft_errors
from PIL.Image import Image as PILImage
from PIL.Image import open as PILImageOpen
from PIL.Image import Resampling
from tqdm import tqdm
from os import mkdir


ROW_BASE_HEIGHT = 28


def verify_inputs(inpaths_str: List[str]) -> List[Path]:
    in_type = None
    ret = []
    for inpath in inpaths_str:
        inp = Path(inpath[0])
        if inp.is_dir():
            in_type = 'dir'
            ret.append(inp)
        if in_type is None:
            raise LookupError(f"Argument --source not provided directory. Got {inp.absolute()}")
        if not inp.exists():
            raise FileNotFoundError(f"Argument --source provided file or directory that does not exist. Got {inp.absolute()}")
    return (ret)


def verify_output(outpath: str) -> Tuple[Workbook, Path]:
    outp = Path(outpath[0])
    if outp.is_dir():
        raise FileNotFoundError(f"Argument --out only accepts a path to open/create an xlsx file. Got {outp.absolute()}")
    if not outp.exists():
        wb = Workbook()
        wb.save(outp)
    else:
        wb = openpyxl.open(outp)
    return wb, outp


def init_excel_sheet(wb: Workbook, outpath: Path):
    wb.create_sheet("Intro", index=0)
    sheet: Worksheet = wb["Intro"]
    cell = sheet.cell(row=1, column=1)
    cell.value = "Each distributor receives it's own sheet"
    cell = sheet.cell(row=2, column=1)
    cell.value = "Walking of equivalent glyphs from one distributor to another happens in intermediary sheets."
    cell = sheet.cell(row=3, column=1)
    cell.value = "Each glyph is represented by the glyph image, the md5 of the glyph components, the rasm group name, the ijam group name, the harakat group name, and the set of names given to this glyph."
    cell = sheet.cell(row=4, column=1)
    cell.value = "Each distributor's glyphs initially begin in the unsorted column and then by hand are sorted into distinct groups based on characteristics"
    wb.save(outpath)


if __name__ == '__main__':

    argp = ArgumentParser()
    argp.add_argument("--source",
                      help="""The directory path for multiple .ttf font files.
                      Can be used multiple times!
                      Searches for .ttf & .otf
                      The allowed structure is ...path/to/<distributor-name>
                      The final element in the path must be the distributor name
                      Searches recursively.""",
                      required=True,
                      action='append',
                      nargs='+'
                      )
    argp.add_argument("--out-dir",
                      help="""The path to the output directory.""",
                      required=True,
                      nargs=1
                      )
    ns = argp.parse_args()

    # Verify Args
    inpaths = verify_inputs(ns.source)
    workbook, out_path = verify_output(ns.out_dir)

    init_excel_sheet(workbook, out_path)

    # NOTE SHOULD BE WRITTEN TO NOT OVERWRITE EXISTING WORK WHEN RUN. MAKE NEW TABS.
    font_files = dict()
    for inpath in inpaths:
        print(inpath)
        font_files[inpath] = list(Path.rglob(inpath,'*.ttf')) + list(Path.rglob(inpath,'*.otf'))
    
    # store output
    bad = [] # path, fontfile, name, reason

    font_files_keys = list(font_files.keys())
    for i, font_files_key in enumerate(font_files_keys):
        font_files_key: Path
        fontfiles: List[Path] = font_files[font_files_key]

        # store output
        all_glyph_rows = dict()

        for j, fontfile in enumerate(fontfiles):
            # udpate
            print(f"path: {font_files_key}\t {i+1}/{len(font_files_keys)}\t fontfile: {fontfile} {j+1}/{len(fontfiles)}")

            # open
            font = None
            with open(fontfile, 'rb') as f:
                font = TTFont(f, lazy=False)
            #font.ensureDecompiled()

            try:
                glyphset = font.getGlyphSet()
            except IndexError as e:
                print(f"ERROR: IndexError: {e}")
                bad.append((font_files_key, fontfile, name, f"{type(e)}:{e}"))
                continue

            for name, glyf in glyphset.glyfTable.glyphs.items():
                if '.notdef' in name:
                    continue
                glyf: Glyph

                try:
                    # get the instructions
                    pen = DecomposingRecordingPen(glyphset)
                    glyf.draw(pen, glyphset.glyfTable)
                    instructions = pen.value
                    # generate md5 hash
                    hash = md5(usedforsecurity=False)
                    hash.update(str(instructions).encode())
                    md5_hex = hash.hexdigest()

                    # generate the image of the glyph
                    if md5_hex not in all_glyph_rows:
                        ftpen = FreeTypePen(glyphSet=glyphset)
                        glyf.draw(ftpen, glyfTable=glyphset.glyfTable)
                        glyf_image = ftpen.image(width=0, height=0, contain=True)
                        if glyf_image.size != (0,0):
                            # resize
                            float_height = glyf_image.size[1]
                            if float_height == 0:
                                float_height = 0.0001
                            hpercent = (ROW_BASE_HEIGHT / float_height)
                            wsize = int((float(glyf_image.size[0]) * float(hpercent)))
                            glyf_image = glyf_image.resize((wsize, ROW_BASE_HEIGHT), Resampling.LANCZOS)

                except ft_errors.FT_Exception as e:
                    bad.append((font_files_key, fontfile, name, f"{type(e)}:{e}"))
                    continue # don't save data

                # save data
                if md5_hex in all_glyph_rows:
                    names_set: set = all_glyph_rows[md5_hex]['names']
                    names_set.add(name)
                    all_glyph_rows[md5_hex]['names'] = names_set
                else:
                    all_glyph_rows[md5_hex] = {
                        'image':glyf_image,
                        'names':set([name])
                    }
        
        # construct the temp directory
        TEMP_DIR = './delete_me_tmp'
        TEMP_DIR_PATH = Path(TEMP_DIR)
        if (not TEMP_DIR_PATH.exists()):
            mkdir(TEMP_DIR)

        # save this distributor's path out to excel sheet
        distributor_name = 'unsorted_'+font_files_key.name
        if distributor_name in workbook:
            del workbook[distributor_name]
        workbook.create_sheet(distributor_name)
        dist_ws_raw = workbook[distributor_name]
        dist_ws_raw['A1'] = 'Image'
        dist_ws_raw['B1'] = 'MD5 Hex'
        dist_ws_raw['C1'] = 'Rasm Group'
        dist_ws_raw['D1'] = 'Ijam Group'
        dist_ws_raw['E1'] = 'Harakat Group'
        dist_ws_raw['F1'] = 'Font Names'
        current_row = 1
        for md5_hash, row_dict in tqdm(all_glyph_rows.items()):
            current_row += 1
            # Set row height
            dist_ws_raw.row_dimensions[current_row].height = ROW_BASE_HEIGHT
            # write PIL image
            pil_img: PILImage = row_dict['image']
            if pil_img.size != (0,0):
                img_name = f'{md5_hash}.png'
                img_path = Path(TEMP_DIR_PATH,Path(img_name))
                pil_img.save(img_path)
                pil_img.close()
                pil_img = PILImageOpen(img_path)
                pil_img.load()
                img = openpyxl.drawing.image.Image(img_path.absolute())
                dist_ws_raw.add_image(img=img, anchor=f'A{current_row}')
            else:
                img_path = None
                dist_ws_raw[f'A{current_row}'] = 'empty-image'
            # write md5 hash
            dist_ws_raw[f'B{current_row}'] = md5_hash
            # write names
            dist_ws_raw[f'F{current_row}'] = list(row_dict['names']).__repr__()
        
        # Save out
        workbook.save(out_path)

    # write out the bad rows
    if 'bad' in workbook:
        del workbook['bad'] # remove the old
    workbook.create_sheet('bad')
    bad_ws = workbook['bad']
    bad_ws['A1'] = 'Path'
    bad_ws['B1'] = 'Fontfile'
    bad_ws['C1'] = 'Name'
    bad_ws['D1'] = 'Reason'
    current_row = 1
    for row in bad:
        current_row += 1

        path: Path = row[0]
        bad_ws[f'A{current_row}'] = path.__repr__()
        fontfile: Path = row[1]
        bad_ws[f'B{current_row}'] = fontfile.__repr__()
        bad_ws[f'C{current_row}'] = row[2]
        bad_ws[f'D{current_row}'] = row[3]

    workbook.save(out_path)

    # Remove the temp directory
    for tmp_img in TEMP_DIR_PATH.glob('*'):
        tmp_img: Path
        tmp_img.unlink(missing_ok=True)
    TEMP_DIR_PATH.rmdir()